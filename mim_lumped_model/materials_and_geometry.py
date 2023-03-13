"""materials_and_geometry.py

    Declaration of Geometry and Materials classes

"""
__all__ = ["Geometry", "Materials"]

import matplotlib.pyplot as plt
import numpy as np
from numpy.polynomial import polynomial as poly
from openpyxl import load_workbook, Workbook, worksheet
from scipy import constants as syc


class Geometry:
    """
    Cross-shaped MIM array geometry

    self.a (float) : cross arm width (m)
    self.b (float) : cross arm length (m)
    self.Λ (float) : cross pattern period (m)
    self.t_metal (float) : cross metal film thickness (m)
    self.t_ins (float) : insulator spacer film thickness (m)
    self.c (float) : constant chosen to take the fringe effect of the capacitance
                     and nonuniform electric field distribution into consideration

    """

    def __init__(
        self, a: float, b: float, Λ: float, t_metal: float, t_ins: float, c: float
    ):
        self.a = a
        self.b = b
        self.Λ = Λ
        self.t_metal = t_metal
        self.t_ins = t_ins
        self.c = c
        self.c_prime = 1 - c


class Materials:
    """
    Metal and insulator material properties

    """

    def __init__(
        self,
        insulator_datafile: str,
        insulator_εr_r_model_order: int,
        insulator_εr_i_model_order: int,
        metal_datafile: str,
        metal_n_model_order: int,
        metal_κ_model_order: int,
        absorbance_spectrum_sample_count: int,
        debug: bool = False,
    ):
        """

        Args:
            insulator_datafile (str): Excel file with insulator material property data
            insulator_εr_r_model_order (int): insulator εr.real polynomial model order
            insulator_εr_i_model_order (int): insulator εr.imag polynomial model order
            metal_datafile (str): Excel file with metal material property data
            metal_n_model_order (int): metal n polynomial model order (default = 3)
            metal_κ_model_order (int): metal κ polynomial model order (default = 4)
            absorbance_spectrum_sample_count (int): number of wavelength samples
                                                    in the absorbance spectrum
            debug (bool): enable/disable plotting of optical data with modeled
                          results, for model validation

        Other class variables:
            insulator_εr_r (np.ndarray): insulator εr.real polynomial model coefficients
            insulator_εr_i (np.ndarray): insulator εr.imag polynomial model coefficients
            insulator_name (str): name of insulator
            metal_n (np.ndarray): metal n polynomial model coefficients
            metal_κ (np.ndarray): metal κ polynomial model coefficients
            metal_name (str): name of metal
            σ (float): metal DC conductivity (1/(ohm * meter))
            τ (float): metal relaxation time (s)
            ω_p (float): metal plasma frequency (rads/s)
            λs (np.ndarray): absorbance spectra wavelength domain (largest common
                             wavelength range between metal and insulator optical
                             data sets loaded from Excel files, where the number of
                             samples = absorbance_spectrum_sample_count)

        """

        # Initialize class instance variables
        self.insulator_datafile: str = insulator_datafile
        self.insulator_εr_r_model_order: int = insulator_εr_r_model_order
        self.insulator_εr_i_model_order: int = insulator_εr_i_model_order
        self.metal_datafile: str = metal_datafile
        self.metal_n_model_order: int = metal_n_model_order
        self.metal_κ_model_order: int = metal_κ_model_order
        self.absorbance_spectrum_sample_count: int = absorbance_spectrum_sample_count
        self.debug: bool = debug

        # Declare other class variable types
        self.insulator_εr_r: np.ndarray
        self.insulator_εr_i: np.ndarray
        self.insulator_name: str
        self.metal_n: np.ndarray
        self.metal_κ: np.ndarray
        self.metal_name: str
        self.σ: float
        self.τ: float
        self.ω_p: float
        self.λs: np.ndarray

        # Load material properties and optical index data for metal and insulator,
        # fit the polynomial models to the optical property data
        (
            self.insulator_εr_r,
            self.insulator_εr_i,
            insulator_λs,
            self.insulator_name,
        ) = self.define_insulator_material_properties()
        (
            self.metal_n,
            self.metal_κ,
            metal_λs,
            self.metal_name,
            self.ω_p,
            self.τ,
            self.σ,
        ) = self.define_metal_material_properties()

        # Determine absorbance spectrum wavelength domain (largest common wavelength
        # range between metal and insulator optical data sets loaded from Excel files)
        λ_min: float = max(metal_λs[0], insulator_λs[0])
        λ_max: float = min(metal_λs[-1], insulator_λs[-1])
        self.λs = np.linspace(
            λ_min * 1e-6, λ_max * 1e-6, self.absorbance_spectrum_sample_count
        )

    def ε_ins(self, λ: float) -> complex:
        """
        Insulator complex relative permittivity at wavelength λ (polynomial model)

        Args:
            λ (float): wavelength (m)

        Returns: εr (complex, unit-less)

        """

        return poly.polyval(λ * 1e6, self.insulator_εr_r) + 1j * poly.polyval(
            λ * 1e6, self.insulator_εr_i
        )

    def δ(self, ω: float) -> float:
        """

        Equation 8: metal skin depth at radial frequency ω (polynomial model)

        Args:
            ω (float): radial frequency (Hz)

        Returns: δ (m)

        """

        λ: float = syc.c * (2 * np.pi / ω)
        κ: float = float(poly.polyval(λ * 1e6, self.metal_κ))

        return λ / (2 * np.pi * κ)

    def define_metal_material_properties(
        self,
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, str, float, float, float]:
        """

        Load metal material properties from Excel file and fit polynomial model
        to optical index data

        NB: the self.debug parameter enables/diables plotting of data & modeling results

        Returns: material properties, polynomial models for n & κ

        """

        # Load ωp and τ metal properties
        wb: Workbook = load_workbook(f"data/{self.metal_datafile}")
        properties_ws: worksheet = wb["properties"]
        metal_name: str = properties_ws["B1"].value
        ω_p: float = properties_ws["B2"].value
        τ: float = properties_ws["B3"].value
        σ: float = syc.epsilon_0 * ω_p**2 * τ

        # Load optical data (λ, n, k)
        n_and_k_ws: worksheet = wb["n_and_k"]
        n_and_k_data: np.ndarray = np.array(
            list(n_and_k_ws.iter_rows(min_row=2, max_col=3, values_only=True))
        ).astype(np.float64)
        wb.close()
        λs: np.ndarray = n_and_k_data[:, 0]

        # Fit polynomial models to n(λ) and κ(λ)
        n_poly, n_stats = poly.polyfit(
            x=λs, y=n_and_k_data[:, 1], deg=self.metal_n_model_order, full=True
        )
        if n_stats[0].size == 0:
            raise ValueError(
                f"metal n.real model order ({self.metal_n_model_order}) is too high!"
            )
        κ_poly, κ_stats = poly.polyfit(
            x=λs, y=n_and_k_data[:, 2], deg=self.metal_κ_model_order, full=True
        )
        if κ_stats[0].size == 0:
            raise ValueError(
                f"metal n.imag model order ({self.metal_κ_model_order}) is too high!"
            )

        # If required, plot data and model estimates for validation of goodness of fit
        if self.debug:
            fig, [ax0, ax1] = plt.subplots(2)
            fig.suptitle(
                f"Model fits to n & k data for {metal_name} ({self.metal_datafile})\n"
                rf"ω$_p$ = 2$\pi$ {ω_p/np.pi:.2e} Hz, τ = {τ:.2e} s"
            )
            ax0.plot(λs, n_and_k_data[:, 1], label="data")
            ax0.plot(λs, poly.polyval(n_and_k_data[:, 0], n_poly), "--", label="model")
            ax0.set(
                title=rf"n (order = {self.metal_n_model_order}, "
                f"e$_{{rms}}$ = {n_stats[0][0]:.2e} RIU)",
                ylabel="RIU",
            )
            ax0.grid()
            ax0.legend()
            ax1.plot(λs, n_and_k_data[:, 2], label="data")
            ax1.plot(λs, poly.polyval(n_and_k_data[:, 0], κ_poly), "--", label="model")
            ax1.set(
                title=rf"κ (order = {self.metal_κ_model_order}, "
                f"e$_{{rms}}$ = {κ_stats[0][0]:.2e} RIU)",
                ylabel="RIU",
                xlabel="Wavelength (μm)",
            )
            ax1.grid()
            ax1.legend()
            plt.tight_layout()
            plt.show()

        return n_poly, κ_poly, λs, metal_name, ω_p, τ, σ

    def define_insulator_material_properties(
        self,
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, str]:
        """

        Load insulator material properties from Excel file and fit polynomial model
        to optical index data

        NB: the self.debug parameter enables/diables plotting of data & modeling results

        Returns: polynomial models for insulator relative permittivity components

        """

        # Load optical data (λ, n, k)
        wb: Workbook = load_workbook(f"data/{self.insulator_datafile}")
        properties_ws: worksheet = wb["properties"]
        insulator_name: str = properties_ws["B1"].value
        n_and_k_ws: worksheet = wb["n_and_k"]
        n_and_k_data: np.ndarray = np.array(
            list(n_and_k_ws.iter_rows(min_row=2, max_col=3, values_only=True))
        ).astype(np.float64)
        wb.close()
        λs: np.ndarray = n_and_k_data[:, 0]
        εr_r: np.ndarray = n_and_k_data[:, 1] ** 2 - n_and_k_data[:, 2] ** 2
        εr_i: np.ndarray = 2 * n_and_k_data[:, 1] * n_and_k_data[:, 2]

        # Fit polynomial models to εr.real(λ) and εr.imag(λ)
        εr_r_poly, εr_r_stats = poly.polyfit(
            x=λs, y=εr_r, deg=self.insulator_εr_r_model_order, full=True
        )
        if εr_r_stats[0].size == 0:
            raise ValueError(
                f"insulator ε_r.real model order ({self.insulator_εr_r_model_order})"
                " is too high!"
            )
        εr_i_poly, εr_i_stats = poly.polyfit(
            x=λs, y=εr_i, deg=self.insulator_εr_i_model_order, full=True
        )
        if εr_i_stats[0].size == 0:
            raise ValueError(
                f"insulator ε_r.imag model order ({self.insulator_εr_i_model_order})"
                " is too high!"
            )

        # If required, plot data and model estimates for validation of goodness of fit
        if self.debug:
            εr_r_modeled: np.ndarray = poly.polyval(λs, εr_r_poly)
            εr_i_modeled: np.ndarray = poly.polyval(λs, εr_i_poly)
            fig, [ax0, ax1] = plt.subplots(2)
            fig.suptitle(
                r"Model fits to ε$_r$ real & imaginary component data"
                f" for {insulator_name}"
                f" ({self.insulator_datafile})"
            )
            ax0.plot(λs, εr_r, label="data")
            ax0.plot(λs, εr_r_modeled, "--", label="model")
            ax0.set(
                title=rf"ε$_r$.real (order = {self.insulator_εr_r_model_order}, "
                f"e$_{{rms}}$ = {εr_r_stats[0][0]:.2e})"
            )
            ax0.legend()
            ax0.grid()
            ax1.plot(λs, εr_i, label="data")
            ax1.plot(λs, εr_i_modeled, "--", label="model")
            ax1.set(
                title=rf"ε$_r$.imag (order = {self.insulator_εr_i_model_order}, "
                f"e$_{{rms}}$ = {εr_i_stats[0][0]:.2e})",
                xlabel="Wavelength (μm)",
            )
            ax1.legend()
            ax1.grid()
            plt.tight_layout()
            plt.show()

        return εr_r_poly, εr_i_poly, λs, insulator_name
