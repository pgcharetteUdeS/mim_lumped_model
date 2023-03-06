"""materials_and_geometry.py

    Declaration of Geometry and Materials classes

"""
__all__ = ["Geometry", "Materials"]

import matplotlib.pyplot as plt
import numpy as np
from numpy.polynomial import polynomial as poly
from openpyxl import load_workbook, Workbook, worksheet
from scipy import constants as syc


class Geometry:
    """
    Cross-shaped MIM array geometry

    self.a (float) : cross arm width (m)
    self.b (float) : cross arm length (m)
    self.Λ (float) : cross pattern period (m)
    self.t_metal (float) : cross metal film thickness (m)
    self.t_ox (float) : oxyde spacer film thickness (m)
    self.c (float) : constant chosen to take the fringe effect of the capacitance
                     and nonuniform electric field distribution into consideration

    """

    def __init__(
        self, a: float, b: float, Λ: float, t_metal: float, t_ox: float, c: float
    ):
        self.a = a
        self.b = b
        self.Λ = Λ
        self.t_metal = t_metal
        self.t_ox = t_ox
        self.c = c
        self.c_prime = 1 - c


class Materials:
    """
    Metal and oxyde material properties

    """

    def __init__(
        self,
        oxyde_datafile: str,
        εr_r_model_order: int,
        εr_i_model_order: int,
        metal_datafile: str,
        n_model_order: int,
        κ_model_order: int,
        absorbance_spectrum_sample_count: int,
        debug: bool = False,
    ):
        """

        Args:
            oxyde_datafile (str): Excel file with oxyde material property data
            εr_r_model_order (int): εr.real polynomial model order (default = 9)
            εr_i_model_order (int): εr.imag polynomial model order (default = 12)
            metal_datafile (str): Excel file with metal material property data
            n_model_order (int): n polynomial model order (default = 3)
            κ_model_order (int): κ polynomial model order (default = 4)
            absorbance_spectrum_sample_count (int): number of wavelength samples
                                                   in the absorbance spectrum
            debug (bool): enable/disable plotting of optical data with modeled
                          results, for model validation

        Other class variables:
            εr_r_ox (np.ndarray): oxyde εr.real polynomial model coefficients
            εr_i_ox (np.ndarray): oxyde εr.imag polynomial model coefficients
            oxyde_name (str): name of oxyde
            n_metal (np.ndarray): metal n polynomial model coefficients
            κ_metal (np.ndarray): metal κ polynomial model coefficients
            metal_name (str): name of metal
            σ (float): metal DC conductivity (1/(ohm * meter))
            τ (float): metal relaxation time (s)
            ω_p (float): metal plasma frequency (Hz)
            λs (np.ndarray): wavelength domain (len = absorbance_spectrum_sample_count)

        """

        # Initialize class instance variables
        self.oxyde_datafile: str = oxyde_datafile
        self.εr_r_model_order: int = εr_r_model_order
        self.εr_i_model_order: int = εr_i_model_order
        self.metal_datafile: str = metal_datafile
        self.n_model_order: int = n_model_order
        self.κ_model_order: int = κ_model_order
        self.absorbance_spectrum_sample_count: int = absorbance_spectrum_sample_count
        self.debug: bool = debug

        # Declare other class variable types
        self.εr_r_ox: np.ndarray
        self.εr_i_ox: np.ndarray
        self.oxyde_name: str
        self.n_metal: np.ndarray
        self.κ_metal: np.ndarray
        self.metal_name: str
        self.σ: float
        self.τ: float
        self.ω_p: float
        self.λs: np.ndarray

        # Load material properties and optical index data for the metal and the oxyde,
        # fit the polynomial models to the optical property data
        (
            self.εr_r_ox,
            self.εr_i_ox,
            λs_oxyde,
            self.oxyde_name,
        ) = self.define_oxyde_material_properties()
        (
            self.n_metal,
            self.κ_metal,
            λs_metal,
            self.metal_name,
            self.ω_p,
            self.τ,
            self.σ,
        ) = self.define_metal_material_properties()

        # Determine absorbance spectrum wavelength domain (largest common wavelength
        # range between the metal and oxyde optical data sets loaded from Excel files)
        λ_min: float = max(λs_metal[0], λs_oxyde[0])
        λ_max: float = min(λs_metal[-1], λs_oxyde[-1])
        self.λs = np.linspace(
            λ_min * 1e-6, λ_max * 1e-6, self.absorbance_spectrum_sample_count
        )

    def ε_ox(self, λ: float) -> complex:
        """
        oxyde complex relative permittivity at wavelength λ (polynomial model)

        Args:
            λ (float): wavelength (m)

        Returns: εr (complex)

        """

        return poly.polyval(λ * 1e6, self.εr_r_ox) + 1j * poly.polyval(
            λ * 1e6, self.εr_i_ox
        )

    def δ(self, ω: float) -> float:
        """

        Equation 8: metal skin depth at radial frequency ω (polynomial model)

        Args:
            ω (float): radial frequency (Hz)

        Returns: δ (m)

        """

        λ: float = syc.c * (2 * np.pi / ω)
        κ: float = float(poly.polyval(λ * 1e6, self.κ_metal))

        return λ / (2 * np.pi * κ)

    def define_metal_material_properties(
        self,
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, str, float, float, float]:
        """

        Load metal material properties from Excel file and fit polynomial model
        to optical index data

        NB: the self.debug parameter enables/diables plotting of data & modeling results

        Returns: material properties, polynomial models for n & κ

        """

        # Load ωp and τ metal properties
        wb: Workbook = load_workbook(self.metal_datafile)
        properties_ws: worksheet = wb["properties"]
        metal_name: str = properties_ws["B1"].value
        ω_p: float = properties_ws["B2"].value
        τ: float = properties_ws["B3"].value
        σ: float = syc.epsilon_0 * ω_p**2 * τ

        # Load optical data (λ, n, k)
        n_and_k_ws: worksheet = wb["n_and_k"]
        n_and_k_data: np.ndarray = np.array(
            list(n_and_k_ws.iter_rows(min_row=2, values_only=True))
        ).astype(np.float64)
        wb.close()
        λs: np.ndarray = n_and_k_data[:, 0]

        # Fit polynomial models to n(λ) and κ(λ), order is determined by trial & error
        n_poly, n_stats = poly.polyfit(
            x=λs, y=n_and_k_data[:, 1], deg=self.n_model_order, full=True
        )
        if n_stats[0].size == 0:
            raise ValueError(
                f"metal n.real model order ({self.n_model_order}) is too high!"
            )
        κ_poly, κ_stats = poly.polyfit(
            x=λs, y=n_and_k_data[:, 2], deg=self.κ_model_order, full=True
        )
        if κ_stats[0].size == 0:
            raise ValueError(
                f"metal n.imag model order ({self.κ_model_order}) is too high!"
            )

        # If required, plot data and model estimates for validation of goodness of fit
        if self.debug:
            fig, [ax0, ax1] = plt.subplots(2)
            fig.suptitle(
                f"Model fits to n (order = {self.n_model_order}) & "
                f"k (order = {self.κ_model_order}) data for {metal_name}\n"
                rf"ω$_p$ = 2$\pi$ {ω_p/np.pi:.2e} Hz, τ = {τ:.2e} s"
            )
            ax0.plot(λs, n_and_k_data[:, 1], label="data")
            ax0.plot(λs, poly.polyval(n_and_k_data[:, 0], n_poly), "--", label="model")
            ax0.set(title=rf"n (e$_{{rms}}$ = {n_stats[0][0]:.2e} RIU)", ylabel="RIU")
            ax0.grid()
            ax0.legend()
            ax1.plot(λs, n_and_k_data[:, 2], label="data")
            ax1.plot(λs, poly.polyval(n_and_k_data[:, 0], κ_poly), "--", label="model")
            ax1.set(
                title=rf"κ (e$_{{rms}}$ = {κ_stats[0][0]:.2e} RIU)",
                ylabel="RIU",
                xlabel="Wavelength (μm)",
            )
            ax1.grid()
            ax1.legend()
            plt.tight_layout()
            plt.show()

        return n_poly, κ_poly, λs, metal_name, ω_p, τ, σ

    def define_oxyde_material_properties(
        self,
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, str]:
        """

        Load oxyde material properties from Excel file and fit polynomial model
        to optical index data

        NB: the self.debug parameter enables/diables plotting of data & modeling results

        Returns: polynomial models for oxyde relative permittivity components

        """

        # Load optical data (λ, n, k) from the Excel file worksheet
        wb: Workbook = load_workbook(self.oxyde_datafile)
        properties_ws: worksheet = wb["properties"]
        oxyde_name: str = properties_ws["B1"].value
        n_and_k_ws: worksheet = wb["n_and_k"]
        n_and_k_data: np.ndarray = np.array(
            list(n_and_k_ws.iter_rows(min_row=2, values_only=True))
        ).astype(np.float64)
        wb.close()
        λs: np.ndarray = n_and_k_data[:, 0]
        εr_r: np.ndarray = n_and_k_data[:, 1] ** 2 - n_and_k_data[:, 2] ** 2
        εr_i: np.ndarray = 2 * n_and_k_data[:, 1] * n_and_k_data[:, 2]

        # Fit polynomial models to εr.real(λ) and εr.imag(λ), order determined
        # by trial and error
        εr_r_poly, εr_r_stats = poly.polyfit(
            x=λs, y=εr_r, deg=self.εr_r_model_order, full=True
        )
        if εr_r_stats[0].size == 0:
            raise ValueError(
                f"oxyde ε_r.real model order ({self.εr_r_model_order}) is too high!"
            )
        εr_i_poly, εr_i_stats = poly.polyfit(
            x=λs, y=εr_i, deg=self.εr_i_model_order, full=True
        )
        if εr_i_stats[0].size == 0:
            raise ValueError(
                f"oxyde ε_r.imag model order ({self.εr_i_model_order}) is too high!"
            )

        # If required, plot data and model estimates for validation of goodness of fit
        if self.debug:
            εr_r_modeled: np.ndarray = poly.polyval(λs, εr_r_poly)
            εr_i_modeled: np.ndarray = poly.polyval(λs, εr_i_poly)
            fig, [ax0, ax1] = plt.subplots(2)
            fig.suptitle(
                rf"Model fits to ε$_r$.real (order = {self.εr_r_model_order}) & "
                rf"ε$_r$.imag (order = {self.εr_i_model_order}) data for {oxyde_name}"
            )
            ax0.plot(λs, εr_r, label="data")
            ax0.plot(λs, εr_r_modeled, "--", label="model")
            ax0.set(title=rf"ε$_r$.real (e$_{{rms}}$ = {εr_r_stats[0][0]:.2e})")
            ax0.legend()
            ax0.grid()
            ax1.plot(λs, εr_i, label="data")
            ax1.plot(λs, εr_i_modeled, "--", label="model")
            ax1.set(
                title=rf"ε$_r$.imag (e$_{{rms}}$ = {εr_i_stats[0][0]:.2e})",
                xlabel="Wavelength (μm)",
            )
            ax1.legend()
            ax1.grid()
            plt.tight_layout()
            plt.show()

        return εr_r_poly, εr_i_poly, λs, oxyde_name
